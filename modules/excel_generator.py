from fileinput import filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import os

class ExcelGenerator:
    """Generates professional Excel files with formatting and charts"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=12)
        self.border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_excel(self, df, filename='output.xlsx', include_charts=True,
                     include_summary=True, apply_formatting=True):
        """Create Excel file with data, formatting, and charts"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create data sheet
        data_sheet = wb.create_sheet('Data', 0)
        self._write_data(data_sheet, df, apply_formatting)
        
        # Create summary sheet if requested
        if include_summary:
            summary_sheet = wb.create_sheet('Summary', 0)
            self._create_summary_sheet(summary_sheet, df)
        
        # Create dashboard sheet with charts if requested
        if include_charts:
            dashboard_sheet = wb.create_sheet('Dashboard', 0 if include_summary else 1)
            self._create_dashboard(dashboard_sheet, df)
        
        # Save workbook
        base_dir = os.getcwd()  # Streamlit-safe
        filepath = os.path.join(base_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def _write_data(self, sheet, df, apply_formatting=True):
        """Write data to sheet with formatting"""
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply formatting
                if apply_formatting:
                    # Header row
                    if r_idx == 1:
                        cell.fill = self.header_fill
                        cell.font = self.header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # Data rows
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                        
                        # Zebra striping
                        if r_idx % 2 == 0:
                            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    cell.border = self.border_style
        
        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, sheet, df):
        """Create summary statistics sheet"""
        # Title
        sheet['A1'] = 'Data Summary Report'
        sheet['A1'].font = Font(bold=True, size=16, color='0066CC')
        
        sheet['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        sheet['A2'].font = Font(italic=True, size=10)
        
        # Basic stats
        row = 4
        sheet[f'A{row}'] = 'Dataset Overview'
        sheet[f'A{row}'].font = Font(bold=True, size=14)
        
        row += 1
        stats = [
            ('Total Records', len(df)),
            ('Total Columns', len(df.columns)),
            ('Numeric Columns', len(df.select_dtypes(include=['number']).columns)),
            ('Text Columns', len(df.select_dtypes(include=['object']).columns)),
            ('Missing Values', df.isnull().sum().sum()),
            ('Duplicate Rows', df.duplicated().sum())
        ]
        
        for label, value in stats:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Numeric columns summary
        row += 2
        sheet[f'A{row}'] = 'Numeric Columns Statistics'
        sheet[f'A{row}'].font = Font(bold=True, size=14)
        
        row += 1
        numeric_cols = df.select_dtypes(include=['number']).columns
        
        if len(numeric_cols) > 0:
            # Headers
            headers = ['Column', 'Count', 'Mean', 'Median', 'Std Dev', 'Min', 'Max']
            for c_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=c_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            # Statistics for each numeric column
            for col in numeric_cols:
                stats = [
                    col,
                    df[col].count(),
                    round(df[col].mean(), 2),
                    round(df[col].median(), 2),
                    round(df[col].std(), 2),
                    round(df[col].min(), 2),
                    round(df[col].max(), 2)
                ]
                
                for c_idx, value in enumerate(stats, 1):
                    cell = sheet.cell(row=row, column=c_idx, value=value)
                    
                    if c_idx > 1:  # Numeric columns
                        cell.number_format = '#,##0.00'
                
                row += 1
        
        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            sheet.column_dimensions[col].width = 15
    
    def _create_dashboard(self, sheet, df):
        """Create dashboard with charts"""
        # Title
        sheet['A1'] = 'Analytics Dashboard'
        sheet['A1'].font = Font(bold=True, size=18, color='0066CC')
        
        chart_row = 3
        
        # Chart 1: Bar chart for numeric columns
        numeric_cols = df.select_dtypes(include=['number']).columns
        
        if len(numeric_cols) > 0:
            # Create data for chart
            col = numeric_cols[0]
            
            # Top 10 values
            if df[col].nunique() < 20:
                chart_data = df.groupby(df.index)[col].sum().head(10)
            else:
                chart_data = df[col].value_counts().head(10)
            
            # Write data to sheet
            data_start_row = chart_row
            sheet.cell(row=data_start_row, column=10, value='Category')
            sheet.cell(row=data_start_row, column=11, value=col)
            
            for idx, (cat, val) in enumerate(chart_data.items(), 1):
                sheet.cell(row=data_start_row + idx, column=10, value=str(cat))
                sheet.cell(row=data_start_row + idx, column=11, value=val)
            
            # Create bar chart
            chart = BarChart()
            chart.title = f'{col} Distribution'
            chart.style = 10
            chart.width = 15
            chart.height = 10
            
            data = Reference(sheet, min_col=11, min_row=data_start_row,
                           max_row=data_start_row + len(chart_data))
            cats = Reference(sheet, min_col=10, min_row=data_start_row + 1,
                           max_row=data_start_row + len(chart_data))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            sheet.add_chart(chart, 'A3')
        
        # Chart 2: Line chart for trends (if date column exists)
        date_cols = df.select_dtypes(include=['datetime64']).columns
        
        if len(date_cols) > 0 and len(numeric_cols) > 0:
            date_col = date_cols[0]
            metric_col = numeric_cols[0]
            
            # Group by date
            trend_data = df.groupby(pd.Grouper(key=date_col, freq='D'))[metric_col].sum().reset_index()
            
            # Write trend data
            trend_start_row = chart_row
            sheet.cell(row=trend_start_row, column=15, value='Date')
            sheet.cell(row=trend_start_row, column=16, value=metric_col)
            
            for idx, (date, val) in enumerate(trend_data.values, 1):
                sheet.cell(row=trend_start_row + idx, column=15, value=date)
                sheet.cell(row=trend_start_row + idx, column=16, value=val)
            
            # Create line chart
            line_chart = LineChart()
            line_chart.title = f'{metric_col} Trend Over Time'
            line_chart.style = 12
            line_chart.width = 15
            line_chart.height = 10
            
            data = Reference(sheet, min_col=16, min_row=trend_start_row,
                           max_row=trend_start_row + len(trend_data))
            dates = Reference(sheet, min_col=15, min_row=trend_start_row + 1,
                            max_row=trend_start_row + len(trend_data))
            
            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(dates)
            
            sheet.add_chart(line_chart, 'A20')
        
        # Chart 3: Pie chart for categorical distribution
        cat_cols = df.select_dtypes(include=['object']).columns
        
        if len(cat_cols) > 0:
            col = cat_cols[0]
            
            # Top 5 categories
            pie_data = df[col].value_counts().head(5)
            
            # Write pie data
            pie_start_row = chart_row
            sheet.cell(row=pie_start_row, column=20, value='Category')
            sheet.cell(row=pie_start_row, column=21, value='Count')
            
            for idx, (cat, val) in enumerate(pie_data.items(), 1):
                sheet.cell(row=pie_start_row + idx, column=20, value=str(cat))
                sheet.cell(row=pie_start_row + idx, column=21, value=val)
            
            # Create pie chart
            pie_chart = PieChart()
            pie_chart.title = f'{col} Distribution'
            pie_chart.width = 12
            pie_chart.height = 10
            
            data = Reference(sheet, min_col=21, min_row=pie_start_row,
                           max_row=pie_start_row + len(pie_data))
            labels = Reference(sheet, min_col=20, min_row=pie_start_row + 1,
                             max_row=pie_start_row + len(pie_data))
            
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            
            sheet.add_chart(pie_chart, 'A37')
        
        # Summary metrics
        summary_row = 55
        
        sheet[f'A{summary_row}'] = 'Key Metrics'
        sheet[f'A{summary_row}'].font = Font(bold=True, size=14)
        
        summary_row += 2
        
        metrics = [
            ('Total Records', len(df)),
            ('Data Completeness', f"{(1 - df.isnull().sum().sum() / df.size) * 100:.1f}%"),
            ('Duplicate Records', df.duplicated().sum())
        ]
        
        for label, value in metrics:
            sheet[f'A{summary_row}'] = label
            sheet[f'B{summary_row}'] = value
            
            sheet[f'A{summary_row}'].font = Font(bold=True)
            sheet[f'A{summary_row}'].fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            sheet[f'B{summary_row}'].fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            
            summary_row += 1
